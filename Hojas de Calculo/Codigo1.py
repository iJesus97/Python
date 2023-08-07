"""from openpyxl import load_workbook

FILE_PATH = 'test.xlsx'
SHEET = 'Hoja 1'

workbook = load_workbook(FILE_PATH, read_only = True)
sheet = workbook[SHEET]

for row in sheet.iter_rows():
	print(row[0].value)
"""
from openpyxl import Workbook, load_workbook
from openpyxl.chart import title

libro = Workbook() #Función principal
Hoja1 = libro.create_sheet("Hoja1") #Crear hojas
Hoja2 = libro.create_sheet("Hoja2")
Hoja0 = libro.create_sheet("Hoja0", 0) #Crear una hoja con posición

print(libro.sheetnames) #Imprimir hojas de un libro

sheet = libro.get_sheet_by_name("Sheet") #Asignar una variable a la hoja por defecto
sheet.title = "HojaX"
#libro.remove(sheet) #Eliminar hojas
print(libro.sheetnames) 
libro.save(filename = "PythonTalk.xlsx")
libro2 = load_workbook("test.xlsx")
escritura = libro2.get_sheet_by_name("Hoja 1")
escritura["B3"] = "Prueba escritura"

libro2.save(filename = "test.xlsx")